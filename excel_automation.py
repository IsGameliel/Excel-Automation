import openpyxl as xl
from openpyxl import chart 
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        discounted_price = cell.value * 0.1
        discounted_price_cell = sheet.cell(row, 4)
        discounted_price_cell.value = discounted_price

    cell_column = sheet.cell(1, 4)
    cell_column.value = 'discounted price'


    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    # saving this data to overwrite the current file
    wb.save(filename)